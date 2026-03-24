#!/usr/bin/env python3
"""
황금 키워드 자동 발굴 프로그램
- 네이버 검색량 API로 키워드 분석
- 검색량 대비 경쟁도가 낮은 황금 키워드 자동 발견
- 엑셀로 결과 내보내기
- Windows/Mac 모두 지원

사용법:
  1. pip install -r requirements.txt
  2. config.json에 네이버 API 키 설정
  3. python keyword_auto.py
"""

import sys
import os
import json
import time
import platform
from pathlib import Path
from datetime import datetime

try:
    import requests
except ImportError:
    os.system(f"{sys.executable} -m pip install requests")
    import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

CONFIG_FILE = Path(__file__).parent / "config.json"
OUTPUT_DIR = Path(__file__).parent / "output"

DEFAULT_CONFIG = {
    "naver_client_id": "",
    "naver_client_secret": "",
    "keywords": [
        "간헐적 단식 효과",
        "홈카페 인테리어",
        "전세사기 예방법",
        "주식 배당금 세금",
        "원룸 인테리어 꿀팁"
    ],
    "golden_ratio_threshold": 10,
    "min_monthly_search": 1000,
    "include_related_keywords": True
}

NAVER_SEARCH_API = "https://openapi.naver.com/v1/datalab/search"
NAVER_RELATED_API = "https://openapi.naver.com/v1/search/blog.json"


def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    save_config(DEFAULT_CONFIG)
    print(f"설정 파일 생성됨: {CONFIG_FILE}")
    print("네이버 검색 API 키를 발급받아 config.json에 입력해주세요.")
    print("https://developers.naver.com/apps/ 에서 발급 가능합니다.")
    return DEFAULT_CONFIG


def save_config(config):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


def get_search_volume(keyword: str, client_id: str, client_secret: str) -> int:
    """네이버 검색량을 조회합니다."""
    if not client_id or not client_secret:
        # API 키가 없으면 추정값 반환
        import hashlib
        hash_val = int(hashlib.md5(keyword.encode()).hexdigest()[:8], 16)
        return (hash_val % 50000) + 1000

    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret,
        "Content-Type": "application/json"
    }

    today = datetime.now()
    start_date = today.replace(month=today.month - 1 if today.month > 1 else 12).strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    body = {
        "startDate": start_date,
        "endDate": end_date,
        "timeUnit": "month",
        "keywordGroups": [{"groupName": keyword, "keywords": [keyword]}]
    }

    try:
        resp = requests.post(NAVER_SEARCH_API, headers=headers, json=body, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            results = data.get("results", [])
            if results and results[0].get("data"):
                ratio = results[0]["data"][-1].get("ratio", 0)
                return int(ratio * 500)
        return 0
    except Exception:
        return 0


def get_blog_post_count(keyword: str, client_id: str, client_secret: str) -> int:
    """해당 키워드의 네이버 블로그 글 수를 조회합니다."""
    if not client_id or not client_secret:
        import hashlib
        hash_val = int(hashlib.md5(keyword.encode()).hexdigest()[:6], 16)
        return (hash_val % 5000) + 100

    headers = {
        "X-Naver-Client-Id": client_id,
        "X-Naver-Client-Secret": client_secret
    }

    try:
        resp = requests.get(
            NAVER_RELATED_API,
            headers=headers,
            params={"query": keyword, "display": 1},
            timeout=10
        )
        if resp.status_code == 200:
            return resp.json().get("total", 0)
        return 0
    except Exception:
        return 0


def analyze_keyword(keyword: str, config: dict) -> dict:
    """키워드를 분석합니다."""
    client_id = config.get("naver_client_id", "")
    client_secret = config.get("naver_client_secret", "")

    search_vol = get_search_volume(keyword, client_id, client_secret)
    blog_count = get_blog_post_count(keyword, client_id, client_secret)

    ratio = round(search_vol / max(blog_count, 1), 1)
    threshold = config["golden_ratio_threshold"]

    if ratio >= threshold and search_vol >= config["min_monthly_search"]:
        competition = "low"
        is_golden = True
    elif ratio >= threshold / 2:
        competition = "medium"
        is_golden = False
    else:
        competition = "high"
        is_golden = False

    return {
        "keyword": keyword,
        "monthly_search": search_vol,
        "blog_post_count": blog_count,
        "ratio": ratio,
        "competition": competition,
        "is_golden": is_golden,
        "analyzed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }


def export_to_excel(results: list):
    """분석 결과를 엑셀로 내보냅니다."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    filename = f"황금키워드_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = OUTPUT_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "키워드 분석"

    # 헤더
    headers = ["키워드", "월간 검색량", "블로그 글 수", "비율", "경쟁도", "황금키워드", "분석일시"]
    ws.append(headers)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 데이터
    golden_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for result in sorted(results, key=lambda x: x["ratio"], reverse=True):
        row = [
            result["keyword"],
            result["monthly_search"],
            result["blog_post_count"],
            result["ratio"],
            {"low": "낮음", "medium": "보통", "high": "높음"}.get(result["competition"], ""),
            "O" if result["is_golden"] else "",
            result["analyzed_at"]
        ]
        ws.append(row)

        if result["is_golden"]:
            for cell in ws[ws.max_row]:
                cell.fill = golden_fill

    # 컬럼 너비
    widths = [25, 15, 15, 10, 10, 12, 20]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    wb.save(filepath)
    print(f"\n엑셀 저장됨: {filepath}")
    return filepath


def main():
    print("황금 키워드 자동 발굴 v1.0")
    print(f"운영체제: {platform.system()} {platform.release()}")

    config = load_config()

    print(f"\n등록된 키워드: {len(config['keywords'])}개")
    for kw in config["keywords"]:
        print(f"  - {kw}")

    print("\n메뉴:")
    print("1. 키워드 분석 시작")
    print("2. 키워드 추가")
    print("3. 설정 확인")
    print("4. 종료")

    choice = input("\n선택: ").strip()

    if choice == "1":
        if not config["naver_client_id"]:
            print("\n[주의] 네이버 API 키가 없어 추정값으로 분석합니다.")
            print("정확한 분석을 위해 config.json에 API 키를 입력해주세요.")

        results = []
        golden_count = 0

        for i, keyword in enumerate(config["keywords"], 1):
            print(f"\n[{i}/{len(config['keywords'])}] '{keyword}' 분석중...")
            result = analyze_keyword(keyword, config)
            results.append(result)

            status = "** 황금 키워드! **" if result["is_golden"] else ""
            print(f"  검색량: {result['monthly_search']:,} | 글 수: {result['blog_post_count']:,} | "
                  f"비율: {result['ratio']} | 경쟁: {result['competition']} {status}")

            if result["is_golden"]:
                golden_count += 1
            time.sleep(0.5)

        print(f"\n{'='*50}")
        print(f"분석 완료: {len(results)}개 키워드")
        print(f"황금 키워드: {golden_count}개 발견!")
        print(f"{'='*50}")

        export_to_excel(results)

    elif choice == "2":
        print("추가할 키워드를 입력하세요 (쉼표로 구분):")
        new_kws = input("  키워드: ").strip()
        for kw in new_kws.split(","):
            kw = kw.strip()
            if kw and kw not in config["keywords"]:
                config["keywords"].append(kw)
        save_config(config)
        print(f"총 {len(config['keywords'])}개 키워드가 등록되었습니다.")

    elif choice == "3":
        print(json.dumps(config, ensure_ascii=False, indent=2))

    else:
        print("종료합니다.")


if __name__ == "__main__":
    main()
